#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_summary.py — 从结构化条目生成 xlsx 分类汇总表 + 一段话总结
条目格式（每行一个）：年度,类别,名称,具体信息,级别,时间
  python gen_summary.py data.csv -o out.xlsx --title "标题" [--summary]
只读输入，输出新文件。条目的数据由盘点/核验环节收集。
"""
import sys, csv, os, argparse
from collections import Counter

def build_xlsx(rows, out, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "成果汇总"
    hdr = ["序号", "年度", "类别", "名称", "具体信息", "级别", "时间"]
    H = PatternFill("solid", fgColor="305496")
    CAT = PatternFill("solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="BFBFBF")
    B = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:G1")
    ws.cell(1, 1, title).font = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    for j, h in enumerate(hdr, 1):
        c = ws.cell(2, j, h)
        c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        c.fill = H; c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = B
    r = 3
    for i, row in enumerate(rows, 1):
        year, cat, name, info, level, time = (row + [""] * 6)[:6]
        for j, v in enumerate([i, year, cat, name, info, level, time], 1):
            c = ws.cell(r, j, v)
            c.font = Font(name="微软雅黑", size=9); c.border = B
            c.alignment = Alignment(horizontal="left", wrap_text=True) if j in (4, 5) else Alignment(horizontal="center", wrap_text=True)
        ws.cell(r, 3).fill = CAT
        ws.row_dimensions[r].height = max(20, 14 * (1 + len(str(name)) // 40))
        r += 1
    for j, w in enumerate([6, 8, 14, 44, 40, 14, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    # 分类统计 sheet
    ws2 = wb.create_sheet("分类统计")
    cnt = Counter(row[1] for row in rows)
    ws2.cell(1, 1, "类别"); ws2.cell(1, 2, "数量")
    for i, (k, v) in enumerate(cnt.most_common(), 2):
        ws2.cell(i, 1, k); ws2.cell(i, 2, v)
    wb.save(out)
    return out

def build_summary(rows, name):
    """一段话总结。需要细分计数时请自行在调用处预聚合，此函数给出通用骨架。"""
    total = len(rows)
    cats = Counter(row[1] for row in rows)
    parts = [f"{name}共{total}项"]
    order = ["论文", "软件著作权", "专利", "立项课题", "教材", "教师获奖", "指导学生获奖", "荣誉称号"]
    for k in order:
        if k in cats:
            parts.append(f"{k}{cats[k]}项")
    for k in cats:
        if k not in order:
            parts.append(f"{k}{cats[k]}项")
    return "，".join(parts) + "。"

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("data")
    ap.add_argument("-o", "--out", default="summary.xlsx")
    ap.add_argument("--title", default="盘点汇总")
    ap.add_argument("--name", default="盘点对象")
    ap.add_argument("--summary", action="store_true")
    args = ap.parse_args()
    rows = list(csv.reader(open(args.data, encoding="utf-8")))
    out = build_xlsx(rows, args.out, args.title)
    print("saved:", out)
    if args.summary:
        print(build_summary(rows, args.name))
